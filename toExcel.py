import json
import logging
import os
import sys
from datetime import datetime
from typing import Dict, List, Optional, Union

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

# Constants
DEFAULT_INPUT_FILES = {
    "kpi_data": "car_kpi_data.json",
    "kpi_values": "car_kpi_values_2024.json"
}
DEFAULT_OUTPUT_FILE = "car_kpi_report_2024.xlsx"
MAIN_SHEET_NAME = "Car KPIs"
DASHBOARD_SHEET_NAME = "Dashboard"
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BORDER_STYLE = Side(border_style="thin", color="000000")
TABLE_BORDER = Border(left=BORDER_STYLE, right=BORDER_STYLE, top=BORDER_STYLE, bottom=BORDER_STYLE)
TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium9", 
    showFirstColumn=False,
    showLastColumn=False, 
    showRowStripes=True, 
    showColumnStripes=False
)
CHART_COLORS = ['4472C4', '5B9BD5', 'ED7D31', '70AD47', 'FFC000', '7030A0', 'C00000']


def load_json_file(filename: str) -> Optional[Dict]:
    """
    Load and parse a JSON file.
    
    Args:
        filename: Path to the JSON file
        
    Returns:
        Dict containing the parsed JSON data or None if an error occurred
    """
    try:
        with open(filename, "r") as json_file:
            data = json.load(json_file)
        logger.info(f"Loaded data from {filename}")
        return data
    except FileNotFoundError:
        logger.error(f"Error: {filename} not found.")
        return None
    except json.JSONDecodeError:
        logger.error(f"Error: {filename} contains invalid JSON.")
        return None
    except Exception as e:
        logger.error(f"Unexpected error loading {filename}: {str(e)}")
        return None


def validate_json_data(kpi_data: Dict, values_data: Dict) -> bool:
    """
    Validate that the JSON data contains the expected structure and fields.
    
    Args:
        kpi_data: Dictionary containing KPI metadata
        values_data: Dictionary containing KPI values for each car
        
    Returns:
        Boolean indicating if the data is valid
    """
    # Check for required keys in kpi_data
    if "top_5_KPIs" not in kpi_data or "top_cars_US_2024" not in kpi_data:
        logger.error("KPI data JSON is missing required keys")
        return False
    
    # Check that KPIs and cars are non-empty lists
    if not isinstance(kpi_data["top_5_KPIs"], list) or not kpi_data["top_5_KPIs"]:
        logger.error("KPI list is empty or not a list")
        return False
    
    if not isinstance(kpi_data["top_cars_US_2024"], list) or not kpi_data["top_cars_US_2024"]:
        logger.error("Car list is empty or not a list")
        return False
    
    # Verify that all cars have data in the values JSON
    for car in kpi_data["top_cars_US_2024"]:
        if car not in values_data:
            logger.error(f"Missing data for car '{car}' in values JSON")
            return False
        
        # Check that each car has all required KPIs
        for kpi in kpi_data["top_5_KPIs"]:
            if kpi not in values_data[car]:
                logger.warning(f"Missing KPI '{kpi}' for car '{car}'. Will use default value of 0.")
    
    logger.info("JSON data validation passed")
    return True


def calculate_averages(cars: List[str], kpis: List[str], values: Dict[str, Dict[str, Union[int, float]]]) -> Dict[str, Union[int, float]]:
    """
    Calculate average values for each KPI across all cars.
    
    Args:
        cars: List of car names
        kpis: List of KPI names
        values: Nested dictionary with KPI values for each car
        
    Returns:
        Dictionary with average value for each KPI
    """
    averages = {}
    
    for kpi in kpis:
        # Extract all available values for this KPI
        kpi_values = [values[car].get(kpi, 0) for car in cars]
        
        # Calculate average, handling potential division by zero
        if kpi_values:
            avg = sum(kpi_values) / len(kpi_values)
            # Round to 2 decimal places
            averages[kpi] = round(avg, 2)
        else:
            averages[kpi] = 0
            
    logger.info(f"Calculated averages for {len(kpis)} KPIs")
    return averages


def create_dataframe(cars: List[str], kpis: List[str], values: Dict[str, Dict[str, Union[int, float]]],
                    averages: Dict[str, Union[int, float]]) -> pd.DataFrame:
    """
    Create a pandas DataFrame containing all car KPI data including averages.
    
    Args:
        cars: List of car names
        kpis: List of KPI names
        values: Nested dictionary with KPI values for each car
        averages: Dictionary with average values for each KPI
        
    Returns:
        Pandas DataFrame with all data structured for Excel export
    """
    # Start with car names as the first column
    table_data = {"Car": cars + ["Average"]}
    
    # Add each KPI as a column
    for kpi in kpis:
        kpi_values = [values[car].get(kpi, 0) for car in cars]
        # Add the average as the last row
        kpi_values.append(averages[kpi])
        table_data[kpi] = kpi_values
    
    # Convert to DataFrame
    df = pd.DataFrame(table_data)
    logger.info(f"Created DataFrame with {len(df)} rows and {len(df.columns)} columns")
    return df


def save_to_excel(df: pd.DataFrame, filename: str, sheet_name: str = MAIN_SHEET_NAME) -> bool:
    """
    Save the DataFrame to an Excel file.
    
    Args:
        df: Pandas DataFrame to save
        filename: Output Excel filename
        sheet_name: Name for the main data sheet
        
    Returns:
        Boolean indicating success or failure
    """
    try:
        # Create a directory for the output file if it doesn't exist
        os.makedirs(os.path.dirname(os.path.abspath(filename)) if os.path.dirname(filename) else '.', exist_ok=True)
        
        # Save DataFrame to Excel
        df.to_excel(filename, index=False, sheet_name=sheet_name)
        logger.info(f"Saved data to Excel file: {filename}")
        return True
    except Exception as e:
        logger.error(f"Error saving to Excel: {str(e)}")
        return False


def format_excel_worksheet(ws):
    """
    Apply formatting to the Excel worksheet.
    
    Args:
        ws: openpyxl worksheet object
    """
    # Get the dimensions
    max_row = ws.max_row
    max_col = ws.max_column
    
    # Format headers
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Apply borders to all cells
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).border = TABLE_BORDER
    
    # Format the Average row
    for col in range(1, max_col + 1):
        cell = ws.cell(row=max_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # Auto-size columns
    for col in range(1, max_col + 1):
        column_letter = get_column_letter(col)
        # Find the max length in the column
        max_length = 0
        for row in range(1, max_row + 1):
            cell_value = str(ws.cell(row=row, column=col).value)
            if len(cell_value) > max_length:
                max_length = len(cell_value)
        # Set column width with some padding
        adjusted_width = max_length + 4
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add table styling
    table_ref = f"A1:{get_column_letter(max_col)}{max_row}"
    tab = Table(displayName="CarKPITable", ref=table_ref)
    tab.tableStyleInfo = TABLE_STYLE
    ws.add_table(tab)
    
    logger.info("Applied formatting to Excel worksheet")


def create_excel_charts(wb, kpis: List[str], include_average: bool = True):
    """
    Create bar charts for each KPI on separate sheets.
    
    Args:
        wb: openpyxl workbook object
        kpis: List of KPI names
        include_average: Whether to include the Average in charts
    """
    ws_data = wb[MAIN_SHEET_NAME]
    
    # Get data dimensions
    max_row = ws_data.max_row
    if not include_average:
        max_row -= 1  # Exclude Average row
    
    # Create chart for each KPI
    for i, kpi in enumerate(kpis, start=2):  # Start at column 2 (B) for KPIs
        # Create sheet name (max 31 chars, no special chars)
        sheet_name = f"{kpi.replace(' ', '_')[:31]}"
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws_chart = wb.create_sheet(sheet_name)
        
        # Create bar chart
        chart = BarChart()
        chart.title = f"{kpi} by Car Model (2024)"
        chart.x_axis.title = "Car Model"
        chart.y_axis.title = kpi
        chart.height = 15
        chart.width = 25
        chart.legend = None  # No legend needed for single series
        
        # Set categories (Car names)
        categories = Reference(ws_data, min_col=1, min_row=2, max_row=max_row)
        
        # Set data values (KPI values)
        values = Reference(ws_data, min_col=i, min_row=2, max_row=max_row)
        
        # Add data to chart
        chart.add_data(values)
        chart.set_categories(categories)
        
        # Apply custom colors to bars (cycle through colors)
        s = chart.series[0]
        s.graphicalProperties.solidFill = CHART_COLORS[i % len(CHART_COLORS)]
        
        # Add data labels
        s.dLbls = DataLabelList()
        s.dLbls.showVal = True
        
        # Position chart on the sheet
        ws_chart.add_chart(chart, "B2")
        
        # Add title and description
        ws_chart.cell(row=1, column=1).value = f"{kpi} Performance Comparison"
        ws_chart.cell(row=1, column=1).font = Font(size=14, bold=True)
        
        logger.info(f"Created chart for KPI: {kpi}")
    
    # Create a dashboard with all charts
    create_dashboard(wb, kpis)


def create_dashboard(wb, kpis: List[str]):
    """
    Create a dashboard sheet with summary information and mini charts.
    
    Args:
        wb: openpyxl workbook object
        kpis: List of KPI names
    """
    if DASHBOARD_SHEET_NAME in wb.sheetnames:
        wb.remove(wb[DASHBOARD_SHEET_NAME])
    
    # Create dashboard as the first sheet
    ws_dashboard = wb.create_sheet(DASHBOARD_SHEET_NAME, 0)
    
    # Add title
    ws_dashboard.cell(row=1, column=1).value = "CAR PERFORMANCE DASHBOARD"
    ws_dashboard.cell(row=1, column=1).font = Font(size=16, bold=True)
    ws_dashboard.cell(row=2, column=1).value = f"Generated on {datetime.now().strftime('%Y-%m-%d')}"
    ws_dashboard.cell(row=2, column=1).font = Font(italic=True)
    
    # Create smaller charts for each KPI
    ws_data = wb[MAIN_SHEET_NAME]
    chart_height = 10
    chart_width = 15
    charts_per_row = 2
    
    # Position charts in a grid
    for i, kpi in enumerate(kpis):
        row_pos = 4 + (i // charts_per_row) * (chart_height + 2)
        col_pos = 1 + (i % charts_per_row) * (chart_width + 1)
        
        # Create a small bar chart
        chart = BarChart()
        chart.title = kpi
        chart.height = chart_height
        chart.width = chart_width
        chart.legend = None
        
        # Get max_row excluding Average
        max_row = ws_data.max_row - 1
        
        # Set categories (Car names)
        categories = Reference(ws_data, min_col=1, min_row=2, max_row=max_row)
        
        # Set data values (KPI values) - find the column index for this KPI
        kpi_col = None
        for col in range(1, ws_data.max_column + 1):
            if ws_data.cell(row=1, column=col).value == kpi:
                kpi_col = col
                break
        
        if kpi_col:
            values = Reference(ws_data, min_col=kpi_col, min_row=2, max_row=max_row)
            
            # Add data to chart
            chart.add_data(values)
            chart.set_categories(categories)
            
            # Apply custom color
            s = chart.series[0]
            s.graphicalProperties.solidFill = CHART_COLORS[i % len(CHART_COLORS)]
            
            # Add data labels
            s.dLbls = DataLabelList()
            s.dLbls.showVal = True
            
            # Position chart on the dashboard
            cell_pos = f"{get_column_letter(col_pos)}{row_pos}"
            ws_dashboard.add_chart(chart, cell_pos)
    
    # Add summary insights below the charts
    summary_row = 4 + ((len(kpis) - 1) // charts_per_row + 1) * (chart_height + 2) + 2
    ws_dashboard.cell(row=summary_row, column=1).value = "SUMMARY INSIGHTS"
    ws_dashboard.cell(row=summary_row, column=1).font = Font(size=14, bold=True)
    
    # Add links to individual KPI sheets
    link_row = summary_row + 2
    ws_dashboard.cell(row=link_row, column=1).value = "Detailed KPI Charts:"
    ws_dashboard.cell(row=link_row, column=1).font = Font(bold=True)
    
    for i, kpi in enumerate(kpis):
        sheet_name = f"{kpi.replace(' ', '_')[:31]}"
        
        # Create a hyperlink to the sheet
        ws_dashboard.cell(row=link_row + 1 + i, column=1).value = kpi
        ws_dashboard.cell(row=link_row + 1 + i, column=1).hyperlink = f"#{sheet_name}!A1"
        ws_dashboard.cell(row=link_row + 1 + i, column=1).font = Font(color="0563C1", underline="single")
    
    logger.info("Created dashboard with summary charts")


def main():
    """Main function to orchestrate the KPI data processing workflow."""
    logger.info("Starting Car KPI Data Processing")
    
    # Step 1: Load JSON files
    car_kpi_data = load_json_file(DEFAULT_INPUT_FILES["kpi_data"])
    car_kpi_values = load_json_file(DEFAULT_INPUT_FILES["kpi_values"])
    
    # Check if files loaded successfully
    if not car_kpi_data or not car_kpi_values:
        logger.error("Cannot proceed without both JSON files.")
        return 1
    
    # Step 2: Validate data
    if not validate_json_data(car_kpi_data, car_kpi_values):
        logger.error("JSON data validation failed. Please check the input files.")
        return 1
    
    # Step 3: Extract data
    kpi_list = car_kpi_data["top_5_KPIs"]
    car_list = car_kpi_data["top_cars_US_2024"]
    logger.info(f"Processing {len(kpi_list)} KPIs for {len(car_list)} cars")
    
    # Step 4: Calculate averages
    average_values = calculate_averages(car_list, kpi_list, car_kpi_values)
    
    # Step 5: Create DataFrame
    df = create_dataframe(car_list, kpi_list, car_kpi_values, average_values)
    
    # Step 6: Save to Excel
    if not save_to_excel(df, DEFAULT_OUTPUT_FILE, MAIN_SHEET_NAME):
        logger.error("Failed to save Excel file. Exiting.")
        return 1
    
    # Step 7: Format Excel and add charts
    try:
        wb = load_workbook(DEFAULT_OUTPUT_FILE)
        ws_main = wb[MAIN_SHEET_NAME]
        
        # Apply formatting
        format_excel_worksheet(ws_main)
        
        # Create charts
        create_excel_charts(wb, kpi_list)
        
        # Save workbook with all formatting and charts
        wb.save(DEFAULT_OUTPUT_FILE)
        logger.info(f"Successfully created and formatted Excel report: {DEFAULT_OUTPUT_FILE}")
        
    except Exception as e:
        logger.error(f"Error during Excel formatting and chart creation: {str(e)}")
        return 1
    
    logger.info("Car KPI Data Processing completed successfully!")
    return 0


if __name__ == "__main__":
    sys.exit(main())
